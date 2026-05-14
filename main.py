# main.py
# =============================================================================
# FAE Copilot – FastAPI
# - Extracción de requisitos (heurística + LLM Azure OpenAI)
# - IF-244 (Excel): validación esquema + export a SharePoint
# - Chat con memoria contextual
# =============================================================================

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.concurrency import run_in_threadpool
from typing import Optional, Dict, Any, List, Literal
from pydantic import BaseModel, ValidationError
from sharepoint_indexer import init_db, index_sharepoint, search_local
from sharepoint_indexer import search_sharepoint
from db import get_escandallo_data

import os
import re
import json
import logging
import pyodbc
import time
from dotenv import load_dotenv
load_dotenv()
# =========================================================
# SharePoint SITE (FIJO)
# =========================================================
SITE_ID = "acfae.sharepoint.com,19eeebbb-811d-4c6c-a37e-7f35daa42f6e,9553b920-40be-4436-a943-f11a07abdb44"
# =========================================================
# Azure OpenAI
# =========================================================
from openai import AzureOpenAI

# =========================================================
# Core FAE (tu módulo)
# =========================================================
from fae_core.requirements import (
    read_doc_to_text,
    extract_basic_requirements
)

# =============================================================================
# SQL Server helpers (ENV + cierre + no bloquear)
# =============================================================================

SQL_SERVER = os.getenv("SQL_SERVER")
SQL_DB = os.getenv("SQL_DB", "FAE_LAB")
SQL_USER = os.getenv("SQL_USER")
SQL_PASSWORD = os.getenv("SQL_PASSWORD")
SQL_TRUST = os.getenv("SQL_TRUST", "yes")  # en prod: Encrypt=yes; TrustServerCertificate=no


def get_sql_connection():
    if not all([SQL_SERVER, SQL_DB, SQL_USER, SQL_PASSWORD]):
        raise RuntimeError("Faltan variables de entorno SQL_SERVER, SQL_DB, SQL_USER, SQL_PASSWORD")
    return pyodbc.connect(
        "DRIVER={ODBC Driver 18 for SQL Server};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DB};"
        f"UID={SQL_USER};"
        f"PWD={SQL_PASSWORD};"
        f"TrustServerCertificate={SQL_TRUST};"
    )


# =========================================================
# SharePoint / Graph + Excel
# =========================================================
import base64
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from io import BytesIO
from msal import ConfidentialClientApplication
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _requests_session() -> requests.Session:
    """
    Session con reintentos y backoff para llamadas a Graph.
    """
    session = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=0.8,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "PUT", "POST"]
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def _get_graph_config():
    tenant = os.getenv("GRAPH_TENANT_ID")
    client_id = os.getenv("GRAPH_CLIENT_ID")
    client_secret = os.getenv("GRAPH_CLIENT_SECRET")

    if not all([tenant, client_id, client_secret]):
        raise RuntimeError(
            "Graph no configurado. Define GRAPH_TENANT_ID/CLIENT_ID/CLIENT_SECRET"
        )
    return tenant, client_id, client_secret

_MSAL_CACHE = {"token": None, "exp": 0.0}

def _get_graph_token() -> str:
    now = time.time()
    if _MSAL_CACHE["token"] and _MSAL_CACHE["exp"] - 60 > now:
        return _MSAL_CACHE["token"]

    tenant, client_id, client_secret = _get_graph_config()
    app = ConfidentialClientApplication(
        client_id=client_id,
        authority=f"https://login.microsoftonline.com/{tenant}",
        client_credential=client_secret,
    )

    token = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )

    if "access_token" not in token:
        raise RuntimeError(f"Graph auth error: {token}")

    _MSAL_CACHE["token"] = token["access_token"]
    _MSAL_CACHE["exp"] = now + float(token.get("expires_in", 3600))
    return _MSAL_CACHE["token"]


def create_folder_in_site(
    site_id: str,
    base_path: str,
    project_name: str,
    token: str
):
    session = _requests_session()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # 1. Obtener drive del site CORRECTO
    drive_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive"
    r = session.get(drive_url, headers=headers, timeout=30)
    r.raise_for_status()
    drive_id = r.json()["id"]

    # 2. Crear carpeta dentro del path exacto
    create_url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        f"/root:/{base_path}:/children"
    )

    payload = {
        "name": project_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail"
    }

    r = session.post(create_url, headers=headers, json=payload, timeout=30)

    if r.status_code == 409:
        return {"status": "exists", "folder": project_name}

    r.raise_for_status()
    return {"status": "created", "folder": project_name}


def _encode_share_link(link: str) -> str:
    """Codifica el sharing link de SharePoint para Graph /shares/u!{token}"""
    b64 = base64.b64encode(link.encode("utf-8")).decode("ascii")
    return b64.replace("+", "-").replace("/", "_").rstrip("=")


def _download_from_share_link(share_link: str, token: str) -> bytes:
    enc = _encode_share_link(share_link)
    url = f"https://graph.microsoft.com/v1.0/shares/u!{enc}/driveItem/content"
    session = _requests_session()
    r = session.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    return r.content


def _get_driveitem_metadata(share_link: str, token: str) -> dict:
    enc = _encode_share_link(share_link)
    url = f"https://graph.microsoft.com/v1.0/shares/u!{enc}/driveItem"
    session = _requests_session()
    r = session.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    r.raise_for_status()
    return r.json()


def _get_real_cell(ws, row, col):
    """
    Si la celda pertenece a un merge, devuelve la celda TOP-LEFT del merge.
    """
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            return ws.cell(merged.min_row, merged.min_col)
    return ws.cell(row, col)


def _safe_write(ws, row, col, value):
    cell = _get_real_cell(ws, row, col)
    cell.value = value
    cell.WrapText = True


import tempfile
import os
from openpyxl import load_workbook
from io import BytesIO

def _fill_requirements_sheet(template_bytes, rows_excel, sheet_name="REQUIREMENTS", ext=".xlsx"):
    
    wb = load_workbook(BytesIO(template_bytes), keep_vba=True)
    ws = wb[sheet_name]

    row_excel = 8

    for r in rows_excel:

        ws.cell(row=row_excel, column=2).value = r.get("Id.", "")
        ws.cell(row=row_excel, column=3).value = r.get("SC (Y/N)", "N")
        ws.cell(row=row_excel, column=4).value = r.get("Description", "")
        ws.cell(row=row_excel, column=5).value = r.get("PRIORITY", "")
        ws.cell(row=row_excel, column=6).value = r.get("CATEGORY", "")
        ws.cell(row=row_excel, column=7).value = r.get("INPUT DATE", "")
        ws.cell(row=row_excel, column=8).value = r.get("FAE Comments", "")
        ws.cell(row=row_excel, column=10).value = r.get("Source document", "")
        ws.cell(row=row_excel, column=11).value = r.get("Objective", "")
        ws.cell(row=row_excel, column=12).value = r.get("Description of status", "")
        ws.cell(row=row_excel, column=13).value = r.get("WBS", "")
        ws.cell(row=row_excel, column=14).value = r.get("Status", "")
        ws.cell(row=row_excel, column=15).value = r.get("Validation criteria", "")
        ws.cell(row=row_excel, column=16).value = r.get("Validation Date", "")

        row_excel += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()




def _upload_to_same_folder(
    meta: dict,
    filename: str,
    content: bytes,
    token: str,
    ext: str = ".xlsx"
) -> dict:
    """
    Sube el archivo al mismo directorio que la plantilla, usando el MIME correcto según extensión.
    """
    drive_id = meta["parentReference"]["driveId"]
    parent_id = meta["parentReference"]["id"]
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_id}:/{filename}:/content"

    ext = (ext or "").lower()
    if ext in (".xlsm", ".xltm"):
        content_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
    else:
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": content_type
    }
    session = _requests_session()
    r = session.put(url, headers=headers, data=content, timeout=60)
    r.raise_for_status()
    return r.json()


# =============================================================================
# Configuración básica
# =============================================================================

# Límite de tamaño de archivo subido (25 MB)
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", 25 * 1024 * 1024))

# CORS (restringe en prod con ALLOW_ORIGINS="https://tu-frontend")
ALLOW_ORIGINS = [o.strip() for o in os.getenv("ALLOW_ORIGINS", "*").split(",")]

# Log
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
log = logging.getLogger("fae-copilot")

# =============================================================================
# Helpers JSON (limpieza/parseo robusto) + prompt IF-244 + mapeo a encabezados
# =============================================================================

def _preclean_llm(texto: str) -> str:
    """
    Limpieza mínima para facilitar json.loads sin tocar strings:
    - Elimina fences ```json ... ```
    - Normaliza comillas curvas a rectas
    - Sustituye True/False/None (Python) por true/false/null (JSON) SOLO fuera de strings
    - Quita BOM/espacios
    """
    s = re.sub(r"```(?:json)?\s*([\s\S]*?)```", r"\1", texto, flags=re.IGNORECASE)
    s = s.replace("“", '"').replace("”", '"').replace("’", "'")

    # Sustitución fuera de strings para True/False/None
    out, in_str, esc = [], False, False
    i = 0
    while i < len(s):
        ch = s[i]
        if in_str:
            out.append(ch)
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            i += 1
            continue
        if ch == '"':
            in_str = True
            out.append(ch)
            i += 1
            continue
        if s.startswith("True", i) and re.match(r"\bTrue\b", s[i:]):
            out.append("true"); i += 4; continue
        if s.startswith("False", i) and re.match(r"\bFalse\b", s[i:]):
            out.append("false"); i += 5; continue
        if s.startswith("None", i) and re.match(r"\bNone\b", s[i:]):
            out.append("null"); i += 4; continue
        out.append(ch); i += 1

    return "".join(out).strip().lstrip("\ufeff")


def _fix_common_json_issues(s: str) -> str:
    """Arreglos inocuos: comas colgantes y dobles."""
    s = re.sub(r",\s*]", "]", s)   # ,] -> ]
    s = re.sub(r",\s*}", "}", s)   # ,} -> }
    s = re.sub(r",\s*,", ",", s)   # ,, -> ,
    return s

def extract_if244_query(text: str):

    text = text.upper()

    match = re.search(r'(IF[\s\-_]?244[\s\-_A-Z0-9]+)', text)

    if not match:
        return None

    query = match.group(1)

    # Normalizar separadores
    query = query.replace("-", " ").replace("_", " ")

    # limpiar palabras basura
    STOPWORDS = {
        "DE", "LA", "EL", "DEL",
        "REFERENCIA", "DOC", "DOCUMENTO",
        "BUSCAME", "DAME", "DESCARGAME"
    }

    parts = query.split()
    cleaned = [p for p in parts if p not in STOPWORDS]

    return " ".join(cleaned)

def _extract_first_json_array(texto: str) -> str:
    """Extrae por balanceo de corchetes el PRIMER array JSON que encuentre."""
    start = texto.find("[")
    if start == -1:
        raise ValueError("No '['")
    depth, in_string, esc = 0, False, False
    for i in range(start, len(texto)):
        ch = texto[i]
        if in_string:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return texto[start:i+1]
    raise ValueError("No matching ']'")

def _extract_first_json_object(texto: str) -> str:
    """Extrae por balanceo de llaves el PRIMER objeto JSON que encuentre."""
    start = texto.find("{")
    if start == -1:
        raise ValueError("No '{'")
    depth, in_string, esc = 0, False, False
    for i in range(start, len(texto)):
        ch = texto[i]
        if in_string:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return texto[start:i+1]
    raise ValueError("No matching '}'")


def extraer_json_llm(texto: str) -> Any:
    """
    Parser robusto:
      1) Busca entre anclas BEGIN_JSON ... END_JSON.
      2) Intenta json.loads del texto completo.
      3) Array balanceado.
      4) Objeto balanceado -> lo envuelve en lista.
      5) Regex del bloque más largo.
    """
    raw = _preclean_llm(texto)

    # 1) Entre anclas
    m = re.search(r"BEGIN_JSON\s*([\s\S]*?)\s*END_JSON", raw, re.IGNORECASE)
    if m:
        candidate = _fix_common_json_issues(m.group(1).strip())
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            pass  # seguimos

    # 2) Intento directo
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # 3) Primer array balanceado
    try:
        arr = _extract_first_json_array(raw)
        return json.loads(_fix_common_json_issues(arr))
    except Exception:
        pass

    # 4) Primer objeto balanceado -> lista
    try:
        obj = _extract_first_json_object(raw)
        return [json.loads(_fix_common_json_issues(obj))]
    except Exception:
        pass

    # 5) Último recurso: regex amplio
    m_arr = re.search(r"\[[\s\S]*]", raw)
    if m_arr:
        return json.loads(_fix_common_json_issues(m_arr.group(0)))
    m_obj = re.search(r"\{[\s\S]*}", raw)
    if m_obj:
        return [json.loads(_fix_common_json_issues(m_obj.group(0)))]

    raise ValueError("No fue posible parsear un JSON válido desde la salida del modelo")
def _extract_standard_reference(text: str) -> str:
    """
    Busca referencias tipo MSL.xx.xx.xxxx, MTS.xx.xx.xxxx, ISO xxxx, etc.
    """
    patterns = [
        r"(MSL\.\d{2}\.\d{2}\.\d{4})",
        r"(MTS\.\d{2}\.\d{2}\.\d{4})",
        r"(ISO\s?\d{3,5})",
        r"(IEC\s?\d{3,5})",
    ]

    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1)

    return ""



def build_prompt_if244_excel(nombre_documento: str, texto_documento: str, mode: str, user_scope: str) -> str:

    return f"""
You are extracting requirements for an automotive IF-244 Requirements Register.

The IF-244 does NOT store sentences.
It stores SHORT TECHNICAL REQUIREMENT TITLES.

From the document, identify every technical requirement and convert it into a short title (3–8 words).

Examples of GOOD IF-244 requirements:

- ISO 9001 certification
- PPAP documentation
- Visual inspection for colour and gloss
- Contamination control measures
- Thermal shock test
- Supply voltage range 6V to 16V

For each requirement, you MUST also classify:

PRIORITY:
- HIGH → certifications, audits, compliance, safety, legal, standards
- MEDIUM → process, control, monitoring, documentation
- LOW → improvements, recommendations

CATEGORY (choose ONLY one):
- QUALITY
- PROCESS
- DOCUMENTATION
- CONTROL
- TESTING
- OTHER

Return ONLY valid JSON between BEGIN_JSON and END_JSON.

BEGIN_JSON
[
  {{
    "id": 1,
    "sc": "N",
    "normative_type": "",
    "description": "short technical requirement title",
    "priority": "HIGH | MEDIUM | LOW",
    "category": "QUALITY | PROCESS | DOCUMENTATION | CONTROL | TESTING | OTHER",
    "input_date": "",
    "fae_comments": "",
    "source_document": "{nombre_documento}",
    "objective": "",
    "status_description": "",
    "wbs": "",
    "status": "",
    "validation_criteria": "",
    "validation_date": ""
  }}
]
END_JSON

DOCUMENT:
{(texto_documento or '')[:12000]}
"""







    return template

from datetime import datetime

def map_rows_for_excel(rows_llm: List[Dict[str, Any]], source_doc: str) -> List[Dict[str, Any]]:

    today = datetime.now().strftime("%d/%m/%Y")
    out: List[Dict[str, Any]] = []

    for r in rows_llm:

        desc = r.get("description", "")
        std_ref = _extract_standard_reference(desc)

        out.append({
            "Id.": r.get("id", ""),

            "SC (Y/N)": (r.get("sc") or "N").strip().upper(),

            "Description": desc,

            "PRIORITY": r.get("priority", "") or "",
            "CATEGORY": r.get("category", "") or "",

            # SIEMPRE lo rellena el FAE
            "INPUT DATE": today,

            "FAE Comments": "",

            # LO QUE FALTABA
            "Source document": std_ref if std_ref else source_doc,


            # TEXTO FIJO DE IF-244
            "Objective": "",


            "Description of status": "",

            # FIJO EN ESTA PLANTILLA
            "WBS": "",

            "Status": "",
            "Validation criteria": "",
            "Validation Date": "",
        })

    return out



# =============================================================================
# Validación IF-244 con Pydantic
# =============================================================================

class If244Row(BaseModel):
    id: int
    sc: Literal["Y","N"] = "N"
    normative_type: Literal["SHALL","MUST","SHOULD","MAY",""] = ""

    description: str = ""
    priority: str = ""
    category: Literal["PRODUCT","PROCESS","QUALITY","COST","DELIVERY","SECURITY","LEGAL","SOFTWARE","OTHER",""] = ""
    input_date: str = ""
    fae_comments: str = ""
    source_document: str
    objective: str = ""
    status_description: str = ""
    wbs: str = ""
    status: Literal["On-Going","Achieved","Not achieved",""] = ""
    validation_criteria: str = ""
    validation_date: str = ""


def _normalize_if244_row(d: Dict[str, Any]) -> Dict[str, Any]:
    d = dict(d or {})

    if d.get("sc"):
        d["sc"] = str(d["sc"]).upper().strip()
    if d.get("category"):
        d["category"] = str(d["category"]).upper().strip()
    if d.get("status"):
        d["status"] = str(d["status"]).strip()
    if not d.get("source_document"):
        d["source_document"] = ""
    return d


def _fix_category(d: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalización 'inteligente' de CATEGORY hacia el set permitido.
    """
    d = dict(d or {})

    if d.get("sc"):
        d["sc"] = str(d["sc"]).upper().strip()

    cat = str(d.get("category", "")).upper().strip()
    if "QUALITY" in cat:
        d["category"] = "QUALITY"
    elif "PROCESS" in cat:
        d["category"] = "PROCESS"
    elif "PRODUCT" in cat:
        d["category"] = "PRODUCT"
    elif "COST" in cat:
        d["category"] = "COST"
    elif "DELIVERY" in cat:
        d["category"] = "DELIVERY"
    elif "SECURITY" in cat:
        d["category"] = "SECURITY"
    elif "LEGAL" in cat:
        d["category"] = "LEGAL"
    elif "SOFTWARE" in cat:
        d["category"] = "SOFTWARE"
    elif cat == "":
        d["category"] = ""
    else:
        d["category"] = "OTHER"

    if d.get("status"):
        d["status"] = str(d["status"]).strip()
    if not d.get("source_document"):
        d["source_document"] = ""
    return d

def _normalize_priority(d: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normaliza cualquier valor de prioridad del LLM al set permitido por el Excel IF-244.
    """
    d = dict(d or {})
    p = str(d.get("priority", "")).lower().strip()

    if "critical" in p:
        d["priority"] = "High"
    elif "high" in p:
        d["priority"] = "High"
    elif "medium" in p:
        d["priority"] = "Medium"
    elif "low" in p:
        d["priority"] = "Low"
    else:
        d["priority"] = ""

    return d



# =============================================================================
# Chunking + inferencia normative_type
# =============================================================================

def _split_text(text: str, chunk: int = 8000, overlap: int = 1000) -> List[str]:
    """Divide texto largo en chunks con solape para evitar cortar requisitos."""
    out: List[str] = []
    i, n = 0, len(text)
    while i < n:
        j = min(n, i + chunk)
        out.append(text[i:j])
        if j == n:
            break
        i = max(0, j - overlap)
    return out

_MODAL_SHALL = re.compile(r'\b(shall|must)\b', re.IGNORECASE)
_MODAL_SHOULD = re.compile(r'\bshould\b', re.IGNORECASE)
_MODAL_MAY = re.compile(r'\bmay\b', re.IGNORECASE)

def _infer_normative_type_if_missing(description: str, current: str) -> str:
    """
    Si el LLM no clasificó el tipo normativo, infiérelo de forma ligera.
    Vacío ("") sigue siendo válido para requisitos no modales/descriptivos.
    """
    cur = (current or "").strip().upper()
    if cur in {"SHALL", "MUST", "SHOULD", "MAY"}:
        return cur
    desc = description or ""
    if _MODAL_SHALL.search(desc):  return "SHALL"
    if _MODAL_SHOULD.search(desc): return "SHOULD"
    if _MODAL_MAY.search(desc):    return "MAY"
    return ""  # No modal: mantener vacío


# =============================================================================
# Configuración Azure OpenAI (variables de entorno)
# =============================================================================
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-11-01-preview")
DEBUG_LLM = os.getenv("DEBUG_LLM", "0") == "1"

client: Optional[AzureOpenAI] = None

# =============================================================================
# Utilidades LLM (centralizar llamada y “reparación”)
# =============================================================================

def _call_llm(prompt: str, temperature: float = 0.0, max_tokens: int = 3000) -> str:
    if not client:
        raise RuntimeError("AzureOpenAI client no inicializado")
    resp = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[
            {"role": "system", "content": "IF-244 requirements engineer. Output JSON only."},
            {"role": "user", "content": prompt},
        ],
        temperature=temperature,
        max_tokens=max_tokens,
    )
    content = resp.choices[0].message.content or ""
    if DEBUG_LLM:
        log.info("[LLM content - first 1200 chars]\n%s", content[:1200])
    return content


def _repair_json_with_llm(raw: str, nombre_documento: str) -> str:
    """
    Segundo intento: pedir al LLM que re-formatee su salida a un JSON válido (array)
    pegado entre BEGIN_JSON/END_JSON.
    """
    repair_prompt = f"""
You will receive your previous output that was NOT valid JSON.
Transform it into a VALID **JSON array** ONLY, enclosed strictly between the markers:

BEGIN_JSON
[ ... JSON array ... ]
END_JSON

Use this schema with concrete values (no alternatives):

[
  {{
    "id": 1,
    "sc": "N",
    "normative_type": "",
    "description": "Verbatim, testable requirement.",
    "priority": "",
    "category": "",
    "input_date": "",
    "fae_comments": "",
    "source_document": "{nombre_documento}",
    "objective": "",
    "status_description": "",
    "wbs": "",
    "status": "",
    "validation_criteria": "",
    "validation_date": ""
  }}
]

Rules:
- Output MUST be ONLY the JSON array between BEGIN_JSON and END_JSON.
- No markdown, no prose, no comments.

Previous output to fix:
{raw[:6000]}
""".strip()
    return _call_llm(repair_prompt, temperature=0.0, max_tokens=1500)

# =============================================================================
# App FastAPI
# =============================================================================

app = FastAPI(
    title="FAE Copilot API",
    version="1.2.0",
    servers=[
        {
            "url": "https://copilot-fae-c0d4huf8hsd7ghbt.westeurope-01.azurewebsites.net",
            "description": "Production"
        }
    ]
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOW_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

# =========================================================
# STARTUP: inicializa Azure OpenAI client y valida envs
# =========================================================
@app.on_event("startup")
def on_startup():
    global client
    if not all([AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_KEY, AZURE_OPENAI_DEPLOYMENT]):
        raise RuntimeError(
            "❌ Faltan variables de entorno de Azure OpenAI "
            "(AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_KEY, AZURE_OPENAI_DEPLOYMENT)"
        )
    client = AzureOpenAI(
        api_key=AZURE_OPENAI_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
    )
    log.info("✅ Azure OpenAI client inicializado")
    #try:
        #_get_graph_config()
    #except Exception:
    #init_db()
        #log.warning("⚠️ Variables de Graph no configuradas (necesarias para export a SharePoint)")

@app.post("/sharepoint/reindex")
async def reindex_sharepoint():
    try:
        from sharepoint_indexer import index_sharepoint

        # 🔥 IMPORTANTE: SIN parámetros extra
        index_sharepoint(
            get_token=_get_graph_token,
            get_session=_requests_session
        )

        return {"status": "ok", "message": "Reindexación completada"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://salmon-rock-07d574203.7.azurestaticapps.net",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"status": "ok"}
# =========================================================
# Redis (opcional en desarrollo)
# =========================================================

import redis
import json
import time

REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379/0")
STATE_TTL_SECONDS = int(os.getenv("STATE_TTL_SECONDS", "86400"))  # 24h

try:
    redis_client = redis.from_url(REDIS_URL, decode_responses=True)
    redis_client.ping()
    REDIS_ENABLED = True
except Exception:
    redis_client = None
    REDIS_ENABLED = False


def _state_key(conversation_id: str) -> str:
    return f"fae:conv:{conversation_id}"


def get_state(conversation_id: str) -> dict:
    if not REDIS_ENABLED or redis_client is None:
        return {}
    raw = redis_client.get(_state_key(conversation_id))
    return json.loads(raw) if raw else {}


def set_state(conversation_id: str, state: dict) -> None:
    if not REDIS_ENABLED or redis_client is None:
        return
    redis_client.set(
        _state_key(conversation_id),
        json.dumps(state),
        ex=STATE_TTL_SECONDS
    )


def set_pending_action(conversation_id: str, skill: str, args: dict, summary: str) -> None:
    state = get_state(conversation_id)
    state["pending_action"] = {
        "skill": skill,
        "args": args,
        "summary": summary,
        "created_at": time.time(),
    }
    set_state(conversation_id, state)


def clear_pending_action(conversation_id: str) -> None:
    state = get_state(conversation_id)
    state.pop("pending_action", None)
    set_state(conversation_id, state)


def store_last_if244(conversation_id: str, rows_ui: list, rows_excel: list, source_doc: str):
    state = get_state(conversation_id)
    state["last_if244_rows_ui"] = rows_ui
    state["last_if244_rows_excel"] = rows_excel
    state["last_if244_source_doc"] = source_doc
    set_state(conversation_id, state)
# =========================================================
# Normalizador sí / no
# =========================================================
YES = {"si", "sí", "s", "ok", "vale", "confirmo", "confirmar"}
NO  = {"no", "n", "cancelar", "cancela"}

def normalize_yes_no(text: str):
    t = (text or "").strip().lower()
    if t in YES:
        return True
    if t in NO:
        return False
    return None
def extract_sharepoint_search_term(name: str):

    name = name.lower()

    # quitar extensiones
    name = name.replace(".xlsx", "")
    name = name.replace(".xlsm", "")
    name = name.replace(".pdf", "")

    parts = name.split("_")

    # devolver primeras partes relevantes
    if len(parts) >= 2:
        return "_".join(parts[:2])

    return parts[0]

def sanitize_filename(name: str) -> str:
    n = (name or "").strip()
    if not n:
        return "IF244.xlsx"
    n = re.sub(r'[\\/:*?"<>|]+', "_", n)
    if not n.lower().endswith(".xlsx"):
        n += ".xlsx"
    return n[:120]

def looks_like_share_link(url: str) -> bool:
    u = (url or "").strip()
    return u.startswith("https://") and ("sharepoint.com" in u or "1drv.ms" in u)

# =========================================================
# Skills (acciones permitidas)
# =========================================================
from pydantic import BaseModel, Field

class CreateFolderArgs(BaseModel):
    site_id: str = Field(..., min_length=10)
    base_path: str = Field(..., min_length=1)
    project_name: str = Field(..., min_length=1)

def skill_sharepoint_create_folder(args: dict) -> dict:
    data = CreateFolderArgs(**args)
    token = _get_graph_token()  # ya existe en tu código [1](https://acfae-my.sharepoint.com/personal/j_nunez_fae_es/Documents/Archivos%20de%20Microsoft%C2%A0Copilot%20Chat/page.tsx)
    return create_folder_in_site(
        site_id=data.site_id,
        base_path=data.base_path,
        project_name=data.project_name,
        token=token,
    )  # ya existe en tu código [1](https://acfae-my.sharepoint.com/personal/j_nunez_fae_es/Documents/Archivos%20de%20Microsoft%C2%A0Copilot%20Chat/page.tsx)

SKILLS = {
    "sharepoint.create_folder": skill_sharepoint_create_folder
}

# =========================================================
# Skill: buscar documento en SharePoint por nombre y devolver URL
# =========================================================
class FindDocumentArgs(BaseModel):
    query: str = Field(..., min_length=2)

def skill_sharepoint_find_document(args: dict):

    data = FindDocumentArgs(**args)
    query = data.query

    token = _get_graph_token()
    session = _requests_session()

    headers = {
        "Authorization": f"Bearer {token}"
    }

    from urllib.parse import quote
    query_encoded = quote(query)

    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drive/root/search(q='{query_encoded}')"

    r = session.get(url, headers=headers, timeout=30)
    r.raise_for_status()

    data_json = r.json()

    results = []
    for item in data_json.get("value", []):
        results.append({
            "name": item.get("name"),
            "url": item.get("webUrl")
        })

    return {"results": results}

    try:
        # 🔥 descargar desde SharePoint (en memoria)
        file_bytes = _download_from_share_link(url, token)

        # 🔥 convertir a texto
        text = read_doc_to_text(
            file_name="document",
            file_bytes=file_bytes
        )

        return {
            "text": text[:12000]  # limitar tamaño
        }

    except Exception as e:
        return {"error": str(e)}

    # =========================================================
    # 🔥 1. INTENTO GLOBAL (search/query)
    # =========================================================
    try:
        url = "https://graph.microsoft.com/v1.0/search/query"

        payload = {
            "requests": [
                {
                    "entityTypes": ["driveItem"],
                    "query": {
                        "queryString": data.query
                    },
                    "from": 0,
                    "size": 25
                }
            ]
        }

        r = session.post(url, headers=headers, json=payload, timeout=30)

        if r.status_code == 200:
            data_json = r.json()

            hits = data_json["value"][0]["hitsContainers"][0]["hits"]

            results = []
            for h in hits:
                resource = h["resource"]
                results.append({
                    "name": resource.get("name"),
                    "url": resource.get("webUrl")
                })

            if results:
                return {"results": results}

    except Exception as e:
        print("⚠️ Global search falló:", e)

    # =========================================================
    # 🔥 2. FALLBACK → TU MÉTODO ACTUAL (SITE_ID)
    # =========================================================
    try:
        from urllib.parse import quote

        query_encoded = quote(data.query)

        url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drive/root/search(q='{query_encoded}')"

        r = session.get(url, headers=headers, timeout=30)
        r.raise_for_status()

        data_json = r.json()

        results = []
        for item in data_json.get("value", []):
            results.append({
                "name": item.get("name"),
                "url": item.get("webUrl")
            })

        return {"results": results}

    except Exception as e:
        print("❌ Fallback también falló:", e)
        return {"results": []}

# Añadir al diccionario de skills
SKILLS["sharepoint.find_document"] = skill_sharepoint_find_document



# =========================================================
# CHAT GENERAL (Azure OpenAI REAL + CONTEXTO AUTOMÁTICO)
# =========================================================
@app.post("/chat")
async def chat(
    message: str = Form(...),
    mode: str = Form("chat"),
    conversation_id: str = Form("default"),
):
    """
    Chat general conectado a Azure OpenAI.
    Usa automáticamente requisitos extraídos previamente si existen en la conversación.
    """
    try:
        # =========================================================
        # 0) Confirmación de acción pendiente (sí/no)
        # =========================================================
        yn = normalize_yes_no(message)

        state = get_state(conversation_id)  # Redis [1](https://acfae-my.sharepoint.com/personal/j_nunez_fae_es/Documents/Archivos%20de%20Microsoft%C2%A0Copilot%20Chat/page.tsx)
        pending = state.get("pending_action")

        if pending and yn is not None:
            if yn is True:
                skill_name = pending["skill"]
                args = pending["args"]

                clear_pending_action(conversation_id)

                if skill_name not in SKILLS:
                    return {"response": "⚠️ La acción pendiente ya no está disponible."}

                result = SKILLS[skill_name](args)

                # Mensaje humano según resultado
                if result.get("status") == "exists":
                    return {
                        "response": f"ℹ️ La carpeta ya existía: \"{args.get('project_name')}\"",
                        "action": skill_name,
                        "result": result,
                    }
                return {
                    "response": f"✅ Carpeta creada: \"{args.get('project_name')}\"",
                    "action": skill_name,
                    "result": result,
                }

                if skill_name == "if244.export":
                    args = pending["args"]
                    share_link = args["share_link"]
                    file_name = sanitize_filename(args["file_name"])
                    sheet_name = args.get("sheet_name", "REQUIREMENTS")

                    state = get_state(conversation_id)
                    rows_excel = state.get("last_if244_rows_excel") or []
                    if not rows_excel:
                        clear_pending_action(conversation_id)
                        return {"response": "⚠️ No tengo requisitos IF‑244 para exportar. Extrae primero en Registros."}

                    token = _get_graph_token()

                    template_bytes = _download_from_share_link(share_link, token)
                    meta = _get_driveitem_metadata(share_link, token)

                    _, template_ext = os.path.splitext(meta.get("name", "IF244_template.xlsx"))
                    template_ext = (template_ext or ".xlsx").lower()

                    filled_bytes = _fill_requirements_sheet(
                        template_bytes,
                        rows_excel,
                        sheet_name=sheet_name,
                        ext=template_ext
                )

                uploaded = _upload_to_same_folder(meta, file_name, filled_bytes, token, ext=template_ext)

                clear_pending_action(conversation_id)

                return {
                    "response": f"✅ Excel generado: **{uploaded.get('name', file_name)}**\n\n{uploaded.get('webUrl','')}",
                    "result": {"fileName": uploaded.get("name", file_name), "webUrl": uploaded.get("webUrl")}
                }

            # yn == False
            clear_pending_action(conversation_id)
            return {"response": "❎ Acción cancelada. No he realizado ningún cambio."}

        if pending and yn is None:
            return {
                "response": f"{pending.get('summary', 'Tienes una acción pendiente.')}\nResponde 'sí' para confirmar o 'no' para cancelar."
            }
        
        # =========================================================
        # INTENT: buscar documento en SharePoint por nombre (wizard)
        # =========================================================
        msg = (message or "").strip()
        import re

        # =========================================================
        # 🔥 DETECCIÓN IF244 (NUEVA)
        # =========================================================
        query_if244 = extract_if244_query(msg)

        if query_if244:

            base_query = query_if244

            # 🔥 diferentes versiones de búsqueda
            queries = []

            # 🔥 1. ORIGINAL (la más importante)
            queries.append(base_query)

            # 🔥 2. formatos típicos
            queries.append(base_query.replace(" ", "_"))
            queries.append(base_query.replace(" ", "-"))

            parts = base_query.split()

            # 🔥 3. combinaciones útiles
            if len(parts) >= 2:
                queries.append(" ".join(parts[:2]))   # IF244 CTS
                queries.append(parts[-1])             # GLOBAL / DOBLEFUNCION

            # 🔥 4. quitar basura tipo "0"
            queries = [
                q for q in queries
                if len(q.strip()) >= 2 and not q.strip().isdigit()
            ]

            # 🔥 añadir versiones reducidas (muy importante)
            parts = base_query.split()

            if len(parts) > 2:
                queries.append(" ".join(parts[:2]))  # IF244 CTS
                queries.append(parts[-1])           # DOBLEFUNCION

            hits = []

            for q in queries:
                print("🔍 BUSCANDO:", q)

                result = SKILLS["sharepoint.find_document"]({"query": q}) or {}

                hits = result.get("results") or []

                print("➡️ RESULTADOS:", len(hits))

                if hits:
                    break  # en cuanto encuentre algo, paramos

            if not hits:
                return {"response": "❌ No he encontrado ningún IF244."}

            # 🔥 APLICAR SCORING (igual que antes)
            query_upper = base_query.upper()

            def score(item):
                name = (item.get("name") or "").upper()
                score = 0

                if query_upper in name:
                    score += 200

                for p in parts:
                    if p in name:
                        score += 50

                if "GLOBAL" in name and "GLOBAL" not in query_upper:
                    score -= 100

                return score

            hits_sorted = sorted(hits, key=score, reverse=True)
            top = hits_sorted[0]

            state = get_state(conversation_id)
            state["last_document_url"] = top["url"]
            state["last_document_name"] = top["name"]
            set_state(conversation_id, state)

            return {
                "response": f"📄 **{top['name']}**\n\n👉 {top['url']}",
                "result": hits_sorted
            }
        # detectar posible nombre de archivo
        file_match = re.search(
            r"(IF[\-\s]?244[\s\-_A-Z0-9\.]+)",
            msg.upper()
        )
        

        if file_match:

            query = file_match.group(0)

            result = SKILLS["sharepoint.find_document"]({"query": query})

            
            hits = result.get("results") or []

            if not hits:
                return {
                    "response": f"❌ No he encontrado ningún documento llamado '{query}' en SharePoint."
                }

            query_upper = query.upper()

            def score(item):
                name = (item.get("name") or "").upper()
                query = query_upper

                score = 0

                # 🔥 1. MATCH EXACTO (MUY IMPORTANTE)
                if name == query:
                    return 1000

                # 🔥 2. MATCH CASI EXACTO
                if query in name:
                    score += 200

                # 🔥 3. PARTES DEL NOMBRE (KEYWORDS)
                parts = re.split(r'[_\-\s]+', query)

                for p in parts:
                    if len(p) > 3 and p in name:
                        score += 50

                # 🔥 4. BONUS SI CONTIENE MUCHAS PARTES
                matches = sum(1 for p in parts if p in name)
                score += matches * 20

                # 🔥 5. PENALIZAR "GLOBAL" si no está en query
                if "GLOBAL" in name and "GLOBAL" not in query:
                    score -= 100

                # 🔥 6. BONUS IF244 solo si lo piden
                if "IF244" in name and "IF244" in query:
                    score += 10
                
                # 🔥 BOOST por keywords importantes
                if "CTS" in name:
                    score += 30

                if "DOBLEFUNCION" in name:
                    score += 30

                return score


            hits_sorted = sorted(hits, key=score, reverse=True)

            top = hits_sorted[0]
            clean_url = top["url"]

            return {
                "response": f"""📄 **{top['name']}**

            👉 [Abrir documento en SharePoint]({top['url']})""",
                    "result": result
            }

        # Si el usuario dice "buscar documento" o "dame el documento" etc.
        if any(k in msg.lower() for k in [
        "buscar",
        "documento",
        "archivo"
        ]):
            pending = {
                "skill": "sharepoint.find_document",
                "args": {},
                "stage": "await_name",
                "summary": "Dime el nombre (o parte del nombre) del documento de SharePoint y te devuelvo la URL."
            }
            state = get_state(conversation_id)
            state["pending_action"] = pending
            set_state(conversation_id, state)
            return {"response": pending["summary"], "pending": True}

        #   Si estamos esperando el nombre
        state = get_state(conversation_id)
        pending = state.get("pending_action") or {}
        if pending.get("skill") == "sharepoint.find_document" and pending.get("stage") == "await_name":
            query = msg
            clear_pending_action(conversation_id)

            result = SKILLS["sharepoint.find_document"]({"query": query})

            
            hits = result.get("results") or []

            if not hits:
                return {"response": f"❌ No he encontrado nada con '{query}'. Prueba con otra palabra o parte del nombre."}

            # Devolver top 1 directamente + sugerir alternativas si hay más
            top = hits[0]
            extra = ""
            if len(hits) > 1:
                extra_list = "\n".join([f"- {h['name']}: {h['url']}" for h in hits[1:]])
                extra = f"\n\nOtros resultados:\n{extra_list}"

            return {"response": f"✅ Aquí tienes el documento:\n\n{top['name']}\n{top['url']}{extra}", "result": result}

            
        # =========================================================
        # INTENT: exportar IF-244 a SharePoint (wizard)
        # =========================================================
        state = get_state(conversation_id)
        pending = state.get("pending_action")

        # 1) Si hay un export pendiente en modo "captura de datos"
        if pending and pending.get("skill") == "if244.export":
            stage = pending.get("stage")
            args = pending.get("args", {})

        # =========================================================
        # IF244 Export Wizard: capturar link y nombre (SEGURO)
        # =========================================================
        state = get_state(conversation_id)
        pending = state.get("pending_action") or {}
        skill = pending.get("skill")    
        stage = pending.get("stage")
        args = pending.get("args", {}) or {}

        if skill == "if244.export":
            if stage == "await_link":
                share_link = message.strip()

                if not looks_like_share_link(share_link):
                    return {
                        "response": "⚠️ Ese enlace no parece un vínculo válido de SharePoint. "
                            "Pega el link completo (https://...)."
                    }

                args["share_link"] = share_link
                pending["args"] = args
                pending["stage"] = "await_name"
                state["pending_action"] = pending
                set_state(conversation_id, state)

                return {
                    "response": "✅ Link recibido. Ahora dime el **nombre del fichero** final (ej: IF244_ProyectoX.xlsx)."
                }

        if stage == "await_name":
            fname = sanitize_filename(message)
            args["file_name"] = fname
            pending["args"] = args
            pending["stage"] = "await_confirm"

            rows_excel = state.get("last_if244_rows_excel") or []
            pending["summary"] = f"Voy a exportar **{len(rows_excel)}** requisitos y guardar como **{fname}**. ¿Confirmas? (sí/no)"


        # Esperamos el nombre del fichero final
        if stage == "await_name":
            fname = sanitize_filename(message)
            args["file_name"] = fname
            pending["args"] = args
            pending["stage"] = "await_confirm"

            rows_excel = state.get("last_if244_rows_excel") or []
            pending["summary"] = f"Voy a exportar **{len(rows_excel)}** requisitos y guardar como **{fname}**. ¿Confirmas? (sí/no)"
            state["pending_action"] = pending
            set_state(conversation_id, state)
            return {"response": pending["summary"], "pending": True}

        # 2) Detectar inicio de exportación
        msg = message.lower()
        if ("export" in msg and "if244" in msg) or ("exportar" in msg and "if244" in msg) or ("excel" in msg and "if244" in msg):
            rows_excel = state.get("last_if244_rows_excel")
            if not rows_excel:
                return {"response": "⚠️ No tengo requisitos IF‑244 en esta conversación. Primero usa el módulo **Registros** para extraerlos."}

            # Intentar capturar link y/o nombre en el mismo mensaje (si el usuario lo pone todo de golpe)
            link_match = re.search(r"(https?://\S+)", message)
            name_match = re.search(r"(?:llamalo|llámalo|nombre|archivo)\s+(.+)$", message, re.IGNORECASE)

            share_link = link_match.group(1) if link_match else None
            file_name = sanitize_filename(name_match.group(1)) if name_match else None

            pending = {
                "skill": "if244.export",
                "args": {
                    "share_link": share_link,
                    "file_name": file_name,
                    "sheet_name": os.getenv("IF244_SHEET_NAME", "REQUIREMENTS"),
                },
                "stage": "await_link" if not share_link else ("await_name" if not file_name else "await_confirm"),
                "summary": ""
            }

            # Guardar estado
            state["pending_action"] = pending
            set_state(conversation_id, state)

            if pending["stage"] == "await_link":
                return {"response": "Pásame el **link de SharePoint** de la plantilla IF‑244 que quieres usar."}
            if pending["stage"] == "await_name":
                return {"response": "Perfecto. Ahora dime el **nombre del fichero** final (ej: IF244_ProyectoX.xlsx)."}
            # await_confirm
            pending["summary"] = f"Voy a exportar **{len(rows_excel)}** requisitos y guardar como **{file_name}**. ¿Confirmas? (sí/no)"
            state["pending_action"] = pending
            set_state(conversation_id, state)
            return {"response": pending["summary"], "pending": True}


        # =========================================================
        # 1) Detectar intención (MVP) para crear carpeta → proponer acción
        # =========================================================
        msg = message.lower()
        if "crear carpeta" in msg and "copilot if244" in msg:
            m = re.search(r"carpeta\s+(.+?)\s+en\s+copilot", message, re.IGNORECASE)
            if not m:
                return {"response": "No he podido leer el nombre. Usa: 'crear carpeta <nombre> en Copilot IF244'."}

            project_name = m.group(1).strip()

            site_id = os.getenv("COPILOT_IF244_SITE_ID", SITE_ID)  # SITE_ID ya existe [1](https://acfae-my.sharepoint.com/personal/j_nunez_fae_es/Documents/Archivos%20de%20Microsoft%C2%A0Copilot%20Chat/page.tsx)
            base_path = os.getenv("COPILOT_IF244_BASE_PATH", "Copilot IF244")

            summary = f"Voy a crear la carpeta \"{project_name}\" en SharePoint (Copilot IF244). ¿Confirmas? (sí/no)"

            set_pending_action(
                conversation_id,
                skill="sharepoint.create_folder",
                args={"site_id": site_id, "base_path": base_path, "project_name": project_name},
                summary=summary
            )

            return {"response": summary, "pending": True}
        # =========================================================
        # INTENT: crear carpeta en Copilot IF244
        # =========================================================
        msg = message.lower()

        if "crear carpeta" in msg and "copilot if244" in msg:
            # Extraer nombre del proyecto (simple pero efectivo)
            m = re.search(r"carpeta\s+(.+?)\s+en\s+copilot", message, re.IGNORECASE)
            if not m:
                raise HTTPException(
                    status_code=400,
                    detail="No se pudo identificar el nombre del proyecto"
                )

            project_name = m.group(1).strip()

            # Ejecutar ACCIÓN REAL
            token = _get_graph_token()
            result = create_folder_in_site(SITE_ID, "Copilot IF244", project_name, token)

            return {
                "response": f"✔️ La carpeta \"{project_name}\" se ha creado correctamente en Copilot IF244.",
                "action": "create_folder",
                "result": result
            }
        state = get_state(conversation_id)
        memory = state.get("memory", [])

        system_prompt = """
        Eres FAE IA, un asistente técnico interno de la empresa.

        Copilot IF244 es un espacio de trabajo corporativo en SharePoint,
        ubicado en la biblioteca de documentos del proyecto.
        NO es un sistema operativo ni una carpeta local.

        Cuando el usuario pida acciones como:
        - crear carpetas
        - exportar documentos
        - generar IF-244
        en Copilot IF244 u otros espacios corporativos:

        NO expliques cómo hacerlo manualmente.
        NO muestres comandos de Windows, Linux o Git.
        NO preguntes qué es Copilot IF244.

        Interpreta la petición como una acción interna,
        ejecuta la operación usando el backend
        y responde confirmando el resultado de forma concisa.
        """.strip()
        if mode == "comparison":
            system_prompt += (
                "\nEspecialízate en comparar ensayos, normas o textos técnicos "
                "y resaltar diferencias clave de forma estructurada."
            )
        elif mode == "image":
            system_prompt += (
                "\nDescribe imágenes técnicas de forma clara, precisa y profesional."
            )

        if memory and memory.get("requirements"):
            reqs = memory["requirements"]
            doc = memory.get("document", "documento")
            reqs_text = "\n".join(f"- [{r.get('Tipo')}] {r.get('Requisito')}" for r in reqs)
            system_prompt += f"""

Contexto previo:
Se han extraído requisitos técnicos del documento "{doc}".

Requisitos extraídos:
{reqs_text}

Cuando el usuario diga “estos requisitos”, “los requisitos” o haga
preguntas de interpretación, se refiere a ESTA lista.
Explícalos, interprétalos o resúmelos sin pedir que los reenvíe.
"""

        response = client.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": message},
            ],
            temperature=0.3,
            max_tokens=900,
        )
        return {"response": response.choices[0].message.content}

    except Exception as e:
        log.exception("❌ ERROR /chat")
        raise HTTPException(status_code=500, detail=f"chat error: {e}")

# =========================================================
# EXTRACCIÓN DE REQUISITOS (heurística básica existente)
# =========================================================
@app.post("/extract-requirements")
async def extract_requirements(
    file: UploadFile = File(...),
    use_ai: Optional[str] = Form("true"),
    conversation_id: str = Form("default")
):
    try:
        if not file:
            raise HTTPException(status_code=400, detail="Sube un documento .pdf/.docx")

        # Límite de tamaño
        file_bytes = await file.read()
        if len(file_bytes) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail=f"El archivo supera el límite permitido ({MAX_UPLOAD_BYTES} bytes)")
        
        text = read_doc_to_text(file_name=file.filename, file_bytes=file_bytes)


        print("TEXT LENGTH:", len(text))


        if not text or not text.strip():
            return JSONResponse(content={"rows": [], "errors": ["Documento sin texto"]}, status_code=200)

        from fae_core.requirements import extract_requirements_semantic

        rows = extract_requirements_semantic(text=text, document=file.filename)


        state = get_state(conversation_id)

        state["memory"] = {
            "requirements": rows,
            "document": file.filename
        }

        set_state(conversation_id, state)

        return JSONResponse(
            content={"rows": rows, "errors": []},
            status_code=200
        )

    except HTTPException:
        raise
    except Exception as e:
        log.exception("❌ ERROR extract-requirements")
        raise HTTPException(status_code=500, detail=f"extract-requirements error: {e}")

# =========================================================
# EXTRACCIÓN IF-244 ajustada al ESQUEMA EXCEL (con validación, reparación y CHUNKING)
# =========================================================
@app.post("/extract-if244-excel")
async def extract_if244_excel(
    file: UploadFile = File(...),
    mode: str = Form("assisted"),
    scope: str = Form(""),
    conversation_id: str = Form("default"),
):
    """
    Devuelve:
      - rows_ui: para tu tabla elegante
      - rows_excel: formato Excel IF-244
    """

    try:
        if not file:
            raise HTTPException(status_code=400, detail="Sube un documento .pdf/.docx")

        file_bytes = await file.read()
        if len(file_bytes) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="Archivo demasiado grande")

        text = read_doc_to_text(file_name=file.filename, file_bytes=file_bytes)
        if not text or not text.strip():
            return {"rows_ui": [], "rows_excel": [], "errors": ["Documento sin texto"]}

        # 🔎 1️⃣ Intentar extracción estructural por ID
        from fae_core.requirements import extract_by_id_blocks

        id_rows = extract_by_id_blocks(text, file.filename)

        if len(id_rows) >= 30:
            print("Detected structured baseline → using ID extraction")

            validated = []
            for idx, r in enumerate(id_rows, start=1):
                validated.append({
                    "id": idx,
                    "sc": "N",
                    "normative_type": "",
                    "description": r.get("Requisito", ""),
                    "priority": "",
                    "category": "",
                    "input_date": "",
                    "fae_comments": "",
                    "source_document": file.filename,
                    "objective": "",
                    "status_description": "",
                    "wbs": "",
                    "status": "",
                    "validation_criteria": "",
                    "validation_date": ""
                })

            rows_ui = [{
                "Documento": file.filename,
                "Requisito": r["description"],
                "Tipo": "",
                "Ambiguo": False
            } for r in validated]

            rows_excel = map_rows_for_excel(validated, source_doc=file.filename)

            return {
                "rows_ui": rows_ui,
                "rows_excel": rows_excel,
                "errors": []
            }

        # 🔎 2️⃣ Si no es baseline → usar IA (tu flujo actual)
        chunks = _split_text(text, chunk=8000, overlap=1000)

        all_rows = []
        errors = []

        for ch_idx, ch in enumerate(chunks, start=1):
            prompt = build_prompt_if244_excel(file.filename, ch, mode, scope)
            content = _call_llm(prompt, temperature=0.0, max_tokens=1600)

            try:
                data = extraer_json_llm(content)
            except Exception:
                try:
                    repair = _repair_json_with_llm(content, file.filename)
                    data = extraer_json_llm(repair)
                except Exception:
                    data = []

            if not isinstance(data, list):
                data = [data]

            for raw in data:
                try:
                    norm = _normalize_priority(
                        _fix_category(
                            _normalize_if244_row(raw)
                        )
                    )

                    if not norm.get("source_document"):
                        norm["source_document"] = file.filename

                    dct = norm
                    dct["normative_type"] = _infer_normative_type_if_missing(
                        dct.get("description", ""), dct.get("normative_type", "")
                    )

                    all_rows.append(dct)

                except Exception as e:
                    errors.append(str(e))

        rows_ui = [{
            "Documento": file.filename,
            "Requisito": d.get("description", ""),
            "Tipo": d.get("normative_type", ""),
            "Ambiguo": False
        } for d in all_rows]

        rows_excel = map_rows_for_excel(all_rows, source_doc=file.filename)

        return {
            "rows_ui": rows_ui,
            "rows_excel": rows_excel,
            "errors": errors
        }

    except Exception as e:
        return {"rows_ui": [], "rows_excel": [], "errors": [str(e)]}



        # DEDUPE por (description, source_document)
        seen = set()
        validated: List[Dict[str, Any]] = []
        for d in all_rows:
            key = (d.get("description", "").strip().lower(), d.get("source_document", "").strip().lower())
            if key in seen:
                continue
            seen.add(key)
            validated.append(d)

                # Renumera IDs secuenciales
        for idx, d in enumerate(validated, start=1):
            d["id"] = idx

        # 4) rows_ui (para la tabla visual)
        rows_ui: List[Dict[str, Any]] = []
        for d in validated:
            tipo = d.get("normative_type", "")
            rows_ui.append({
                "Documento": file.filename,
                "Requisito": d.get("description", "") or "",
                "Tipo": tipo,
                "Ambiguo": False
            })

        # 5) rows_excel (encabezados EXACTOS Excel)
        rows_excel = map_rows_for_excel(validated, source_doc=file.filename)

        return {
            "rows_ui": rows_ui,
            "rows_excel": rows_excel,
            "errors": errors,
        }


    except Exception as e:
        err = f"extract-if244-excel error: {e}"
        log.exception("❌ %s", err)
        return {"count": 0, "rows_ui": [], "rows_excel": [], "errors": [err]}


# =========================================================
# EXPORT A SHAREPOINT: descarga plantilla por vínculo, rellena y sube
# =========================================================
@app.post("/export-if244-to-sharepoint")
async def export_if244_to_sharepoint(
    share_link: str = Form(...),          # vínculo de tu plantilla en SharePoint
    project_name: str = Form("PROYECTO"), # para nombrar el archivo final
    rows_excel_json: str = Form(...),     # pega aquí extract-if244-excel["rows_excel"]
    sheet_name: str = Form("REQUIREMENTS")
):
    """
    Descarga la plantilla desde share_link, rellena la hoja 'sheet_name' con rows_excel y sube el archivo a la misma carpeta.
    Devuelve el 'webUrl' del archivo subido.
    """
    try:
        # Validación de configuración (lanza si falta)
        _get_graph_config()

        try:
            rows_excel = json.loads(rows_excel_json)
        except Exception:
            raise HTTPException(status_code=400, detail="rows_excel_json no es un JSON válido")

        if not isinstance(rows_excel, list) or not rows_excel:
            raise HTTPException(status_code=400, detail="rows_excel vacío o inválido")

        token = _get_graph_token()

        # 1) Descargar plantilla (la del link que me pasaste)
        template_bytes = _download_from_share_link(share_link, token)

        # 2) Detectar extensión y rellenar preservando macros si procede
        meta = _get_driveitem_metadata(share_link, token)
        template_name = meta.get("name", "IF244_template.xlsx")
        _, template_ext = os.path.splitext(template_name)
        template_ext = (template_ext or ".xlsx").lower()

        filled_bytes = _fill_requirements_sheet(
            template_bytes,
            rows_excel,
            sheet_name=sheet_name,
            ext=template_ext
        )

        # 3) Subir al mismo folder respetando extensión/MIME
        from datetime import datetime
        today = datetime.now().strftime("%Y%m%d")
        safe_proj = re.sub(r'[^A-Za-z0-9_.-]+', '_', project_name).strip('_')
        base_name = f"IF244_{safe_proj}_{today}{template_ext}"

        uploaded = _upload_to_same_folder(meta, base_name, filled_bytes, token, ext=template_ext)

        return {
            "fileName": uploaded.get("name", base_name),
            "webUrl": uploaded.get("webUrl"),
        }

    except HTTPException:
        raise
    except Exception as e:
        log.exception("❌ export-if244-to-sharepoint error")
        raise HTTPException(status_code=500, detail=f"export-if244-to-sharepoint error: {e}")

@app.post("/sharepoint/create-project-folder")
async def create_project_folder(
    project_name: str = Form(...),
    site_id: str = Form(...),
    base_path: str = Form(...)
):
    try:
        token = _get_graph_token()
        return create_folder_in_site(
            site_id=site_id,
            base_path=base_path,
            project_name=project_name,
            token=token
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =========================================================
# SQL TEST  <<< AÑADIDO (cierre + no bloquear)
# =========================================================
@app.get("/sql-test")
async def sql_test():
    def _do():
        conn = get_sql_connection()
        try:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT GETDATE()")
                row = cursor.fetchone()
                return str(row[0]) if row else None
            finally:
                try:
                    cursor.close()
                except Exception:
                    pass
        finally:
            try:
                conn.close()
            except Exception:
                pass

    sql_time = await run_in_threadpool(_do)
    return {"sql_time": sql_time}

# =========================================================
# HEALTHCHECK
# =========================================================
@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/sharepoint/debug-find-site")
async def debug_find_site(
    search: str = "FAE"
):
    try:
        token = _get_graph_token()
        session = _requests_session()
        headers = {"Authorization": f"Bearer {token}"}

        url = f"https://graph.microsoft.com/v1.0/sites?search={search}"
        r = session.get(url, headers=headers, timeout=30)
        r.raise_for_status()

        return r.json()

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

from urllib.parse import quote

@app.get("/sharepoint/search")
def search_sharepoint_files(query: str):
    try:
        token = _get_graph_token()
        session = _requests_session()

        headers = {
            "Authorization": f"Bearer {token}"
        }

        query_encoded = quote(query)

        url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drive/root/search(q='{query_encoded}')"

        r = session.get(url, headers=headers, timeout=60)
        r.raise_for_status()

        data = r.json()

        results = []

        for item in data.get("value", []):
            results.append({
                "name": item.get("name"),
                "url": item.get("webUrl")
            })

        return {"results": results}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug-db")
def debug_db():
    import sqlite3

    conn = sqlite3.connect("sharepoint_index.db")
    cur = conn.cursor()

    cur.execute("SELECT name FROM files LIMIT 50")
    rows = cur.fetchall()

    conn.close()

    return {"files": rows}

# =========================================================
# ESCANDALLO GENERATOR TEST (CSV → estructura)
# =========================================================

from fastapi import UploadFile, File, HTTPException
import shutil
import os


def build_escandallo(df):
    import pandas as pd

    # Tipos
    df["Codigo Componente"] = df["Codigo Componente"].astype(str)
    df["Aplicado en"] = df["Aplicado en"].astype(str)

    # Cantidad
    df["Coeficiente"] = pd.to_numeric(df["Coeficiente"], errors="coerce").fillna(1)

    # Precio (coma → punto)
    df["Precio Unitario"] = (
        df["Precio Unitario"]
        .astype(str)
        .str.replace(",", ".", regex=False)
    )
    df["Precio Unitario"] = pd.to_numeric(df["Precio Unitario"], errors="coerce").fillna(0)

    # Estructura padre-hijo
    edges = df[[
        "Aplicado en",
        "Codigo Componente",
        "Coeficiente",
        "Descripcion Componente",
        "Precio Unitario",
        "Fase",
        "Seccion"
    ]]

    escandallo = {}

    for _, row in edges.iterrows():
        parent = row["Aplicado en"]
        child = row["Codigo Componente"]

        if parent not in escandallo:
            escandallo[parent] = []

        escandallo[parent].append({
            "child": child,
            "desc": row["Descripcion Componente"],
            "qty": float(row["Coeficiente"]),
            "price": float(row["Precio Unitario"]),
            "fase": str(row["Fase"]),
            "seccion": str(row["Seccion"])
        })

    return escandallo

def calculate_cost(root, esc):
    def dfs(node):
        total = 0

        if node not in esc:
            return 0

        for child in esc[node]:
            child_cost = child["price"] * child["qty"]
            sub_cost = dfs(child["child"])

            total += child_cost + sub_cost

        return total

    return dfs(root)

def flatten_escandallo(root, esc):
    rows = []

    def dfs(node, level):

        if node not in esc:
            return

        for child in esc[node]:

            # 🔥 limpiar descripción REAL
            desc = child["desc"]

            if "-" in desc:
                desc = desc.split("-")[-1].strip()

            desc = desc.upper()

            # 🔥 tipo material MEJORADO
            tipo = "Purchased component"
            if "KG" in desc or child["price"] > 4:
                tipo = "Raw material"

            row = {
                "parent": node,
                "level": level + 1,
                "Codigo": child["child"],
                "Descripcion": desc,
                "Cantidad": float(child["qty"]),
                "Precio": float(child["price"]),
                "Tipo": tipo,
                "Fase": child["fase"],
                "Seccion": child["seccion"]
            }

            rows.append(row)

            # 🔥 CLAVE: bajar nivel correctamente
            dfs(child["child"], level + 1)

    dfs(root, 0)
    return rows

def generate_excel_from_flat(template_path, output_path, rows):
    from openpyxl import load_workbook

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["1.1. Bill of materials"]

    start_row = 10

    for i, r in enumerate(rows):
        row_excel = start_row + i

        # Sr No
        ws.cell(row=row_excel, column=2).value = i + 1

        # Nivel BOM
        ws.cell(row=row_excel, column=3).value = r["parent"]

        # Cantidad
        ws.cell(row=row_excel, column=4).value = r["Cantidad"]

        # Código
        ws.cell(row=row_excel, column=5).value = r["Codigo"]

        # Descripción con indent
        indent = "   " * (r["level"] - 1)
        ws.cell(row=row_excel, column=6).value = indent + r["Descripcion"]

        # Tipo material
        ws.cell(row=row_excel, column=7).value = r["Tipo"]

        # UOM
        ws.cell(row=row_excel, column=13).value = "PCS"

        # Precio
        ws.cell(row=row_excel, column=14).value = r["Precio"]

    wb.save(output_path)
    return output_path

def fill_working_plan(template_path, rows):
    from openpyxl import load_workbook

    wb = load_workbook(template_path, keep_vba=True)

    ws = None
    for s in wb.sheetnames:
        if "working" in s.lower():
            ws = wb[s]
            break

    if ws is None:
        raise Exception("No se encontró Working Plan")

    start_row = 6
    row_excel = start_row

    fases_vistas = set()

    for r in rows:

        # 🔥 PROTECCIÓN
        fase = int(r.get("Fase", 0))
        seccion = r.get("Seccion", "")

        if fase == 0:
            continue

        if fase in fases_vistas:
            continue

        fases_vistas.add(fase)

        ws.cell(row=row_excel, column=1).value = fase
        ws.cell(row=row_excel, column=3).value = f"{seccion} - Fase {fase}"

        row_excel += 1

    wb.save(template_path)

   


@app.post("/generate-escandallo-test")
async def generate_escandallo_test(
    catalog: UploadFile = File(...)
):
    try:
        import pandas as pd
        import numpy as np
        import math

        os.makedirs("temp", exist_ok=True)
        path = f"temp/{catalog.filename}"

        # Guardar archivo
        with open(path, "wb") as buffer:
            shutil.copyfileobj(catalog.file, buffer)

        # =========================================================
        # 🔥 LECTURA CSV ULTRA ROBUSTA
        # =========================================================
        try:
            df = pd.read_csv(path, sep=";", encoding="utf-8", engine="python")
        except:
            try:
                df = pd.read_csv(path, sep=";", encoding="latin-1", engine="python")
            except:
                try:
                    df = pd.read_csv(path, sep=",", encoding="utf-8", engine="python")
                except:
                    df = pd.read_csv(path, sep=None, engine="python")

        # =========================================================
        # 🔥 LIMPIEZA
        # =========================================================
        df = df.replace([np.inf, -np.inf], None)
        df = df.fillna("")

        print("COLUMNS:", df.columns.tolist())
        print("SHAPE:", df.shape)

        # =========================================================
        # 🔥 BUILD ESCANDALLO
        # =========================================================
        esc = build_escandallo(df)
        root = str(df["Acabado"].iloc[0])
        coste_total = calculate_cost(root, esc)

        # =========================================================
        # 🔥 SANITIZAR JSON (CLAVE)
        # =========================================================
        def clean(obj):
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return 0
                return obj
            elif isinstance(obj, dict):
                return {k: clean(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean(v) for v in obj]
            return obj

        esc = clean(esc)

        # 🔥 GENERAR EXCEL
        flat = flatten_escandallo(root, esc)

        template_path = "template.xlsm"
        output_path = f"temp/escandallo_{root}.xlsm"

        generate_excel_from_flat(template_path, output_path, flat)
        fill_working_plan(output_path, flat)

        # =========================================================
        # 🔥 RESPUESTA FINAL
        # =========================================================
        from fastapi.responses import FileResponse

        return FileResponse(
            path=output_path,
            filename=f"escandallo_{root}.xlsm",
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# =========================================================
# ESCANDALLO GENERATOR (NUEVO - DESDE BD)
# =========================================================

@app.get("/escandallo/{ref}")
async def escandallo_from_db(ref: str):
    import shutil
    from fastapi.responses import FileResponse
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side
    import pandas as pd

    def _generate():
        conn = get_sql_connection()
        codigo_concat = "CONCAT(a.n0,a.n1,a.n2,a.n3,a.n4,a.n5,a.n6,a.n7,a.n8,a.n9)"

        # --- BOM ---
        df_bom = pd.read_sql(f"""
            SELECT a.NIvel_Superior AS nivel_bom,
                   {codigo_concat} AS fae_pn,
                   a.coef          AS cantidad,
                   s.ItemName      AS descripcion,
                   a.Tipo_articulo AS tipo,
                   a.PMP           AS pmp,
                   f.Fase          AS fase,
                   f.Descripcion   AS desc_fase,
                   f.Seccion       AS seccion
            FROM dbo.PP_TB_EST_EXT_PMP a
            LEFT JOIN dbo.PP_TB_FASES f ON a.id_fase = f.id_fase
            LEFT JOIN dbo.[SAP_OITM - Items_v00] s
                ON {codigo_concat} = s.ItemCode
            WHERE a.Acabado = '{ref}' AND a.Tipo = 'Actual' AND a.n0 = ''
            ORDER BY a.id ASC
        """, conn)

        # --- Working Plan ---
        df_wp = pd.read_sql(f"""
            WITH fases_raw AS (
                SELECT DISTINCT f.Fase, f.Descripcion, f.Seccion,
                       CASE WHEN UPPER(f.Seccion) = 'ZERVE3' THEN 2 ELSE 1 END AS prioridad
                FROM dbo.PP_TB_EST_EXT_PMP a
                LEFT JOIN dbo.PP_TB_FASES f ON a.id_fase = f.id_fase
                WHERE a.Acabado = '{ref}' AND a.Tipo = 'Actual'
                  AND a.n0 = '' AND f.Fase IS NOT NULL
            ),
            ranked AS (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY Fase ORDER BY prioridad) AS rn
                FROM fases_raw
            )
            SELECT Fase, Descripcion, Seccion FROM ranked WHERE rn = 1 ORDER BY Fase
        """, conn)
        conn.close()
        return df_bom, df_wp

    try:
        df_bom, df_wp = await run_in_threadpool(_generate)

        if df_bom.empty:
            raise HTTPException(status_code=404, detail=f"Referencia {ref} no encontrada")

        # --- Generar xlsm ---
        TIPO_MAP = {
            "Componente": "Purchased component", "Subconjunto": "Sub-assembly",
            "BULK": "Sub-assembly", "Artículo Comprado": "Purchased component",
            "Acabado Bulk": "Sub-assembly", "Producto Terminado": "Finished good",
        }
        def map_tipo(tipo, seccion):
            if seccion and isinstance(seccion, str) and seccion.strip().upper() == "ZERVE3":
                return "Raw material"
            return TIPO_MAP.get(tipo, "Purchased component")

        os.makedirs("temp", exist_ok=True)
        output_path = f"temp/escandallo_{ref}.xlsm"
        shutil.copy2("template.xlsm", output_path)

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = load_workbook(output_path, keep_vba=True)

        # 1.1 Bill of materials
        ws_bom = wb["1.1. Bill of materials"]
        for idx, row in enumerate(df_bom.itertuples(index=False), start=0):
            r = 5 + idx
            ws_bom[f"B{r}"] = idx + 1
            ws_bom[f"C{r}"] = row.nivel_bom or ""
            ws_bom[f"D{r}"] = float(row.cantidad) if row.cantidad is not None else 0
            ws_bom[f"E{r}"] = str(row.fae_pn)
            ws_bom[f"F{r}"] = row.descripcion or ""
            ws_bom[f"G{r}"] = map_tipo(row.tipo, row.seccion)
            for col in ["B","C","D","E","F","G"]:
                ws_bom[f"{col}{r}"].border = border
                ws_bom[f"{col}{r}"].alignment = Alignment(vertical="center")

        # 1.2 Working plan — limpiar plantilla y escribir
        ws_wp = wb["1.2. Working plan"]
        for r in range(6, 20):
            for col in ["B","C","D","E","F","G","H","I","J","K","L"]:
                ws_wp[f"{col}{r}"] = None
        for idx, row in enumerate(df_wp.itertuples(index=False), start=0):
            r = 6 + idx
            ws_wp[f"B{r}"] = int(row.Fase) if row.Fase else ""
            ws_wp[f"D{r}"] = row.Descripcion or ""
            ws_wp[f"L{r}"] = row.Seccion or ""
            for col in ["B","D","L"]:
                ws_wp[f"{col}{r}"].border = border
                ws_wp[f"{col}{r}"].alignment = Alignment(vertical="center")

        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename=f"escandallo_{ref}.xlsm",
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    except HTTPException:
        raise
    except Exception as e:
        log.exception("❌ ERROR escandallo_from_db")
        raise HTTPException(status_code=500, detail=str(e))