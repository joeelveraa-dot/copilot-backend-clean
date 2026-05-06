import openpyxl


def write_bom_to_excel(bom_df, template):

    wb = openpyxl.load_workbook(template, keep_vba=True)

    ws = wb["1.1. Bill of materials"]

    start_row = 5

    for i, row in bom_df.iterrows():

        r = start_row + i

        ws.cell(row=r, column=2).value = i + 1
        ws.cell(row=r, column=3).value = row["parent"]

        cell = ws.cell(row=r, column=4)
        cell.value = row["quantity"]
        cell.number_format = "0.00000"

        ws.cell(row=r, column=5).value = row["part_number"]
        # Item Description
        ws.cell(row=r, column=6).value = row.get("Item Description")

        # Material Type
        ws.cell(row=r, column=7).value = row.get("Material type")

    output_file = "temp/escandallo_generado.xlsm"

    wb.save(output_file)

    return output_file